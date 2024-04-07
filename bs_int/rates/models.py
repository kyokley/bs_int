import csv
import logging
import math
from collections import namedtuple
from io import BytesIO, StringIO
from pathlib import Path

import matplotlib.pyplot as plt
import requests
from dateutil.parser import parse
from django.core.files.base import ContentFile
from django.db import models
from django.utils import timezone
from django.utils.html import mark_safe
from openpyxl import load_workbook
from openpyxl.chart import Reference, ScatterChart, Series

REQUEST_TIMEOUT = 60

CURRENT_MONTH_TREASURY_URL_TEMPLATE = "https://home.treasury.gov/resource-center/data-chart-center/interest-rates/daily-treasury-rates.csv/all/{year}{month}?type=daily_treasury_yield_curve&field_tdr_date_value_month={year}{month}&page&_format=csv"

ANNUAL_TREASURY_URL_TEMPLATE = "https://home.treasury.gov/resource-center/data-chart-center/interest-rates/daily-treasury-rates.csv/{year}/all?type=daily_treasury_yield_curve&field_tdr_date_value={year}&page&_format=csv"

logger = logging.getLogger(__file__)

EXCEL_TEMPLATE_FILE = Path(__file__).parent / "quant_assessment_template.xlsx"
PAR_VALUES_COL = "C"
PAR_VALUES_START_ROW = 19
ZERO_RATES_COL = "E"
ZERO_RATES_START_ROW = 19

MINI_CHART_MONTH_COL_INDEX = 2
CHART_MONTH_COL_INDEX = 10
CHART_PAR_COL = 3
CHART_ZERO_COL = 12
CHART_MIN_ROW = 3
CHART_MAX_ROW = 351
MINI_CHART_MIN_ROW = 19
MINI_CHART_MAX_ROW = 26

Maturity = namedtuple("Maturity", "name,months")
DataRow = namedtuple("DataRow", "months,par,zero,zero_rate,df,ln_df")


class TreasuryData(models.Model):
    _treasury_map = {
        "1 Mo": "one_month",
        "1 Yr": "one_year",
        "10 Yr": "ten_year",
        "2 Mo": "two_month",
        "2 Yr": "two_year",
        "20 Yr": "twenty_year",
        "3 Mo": "three_month",
        "3 Yr": "three_year",
        "30 Yr": "thirty_year",
        "4 Mo": "four_month",
        "5 Yr": "five_year",
        "6 Mo": "six_month",
        "7 Yr": "seven_year",
    }

    _maturity_order = (
        # Maturity(name='six_month', months=6),
        Maturity(name="one_year", months=12),
        Maturity(name="two_year", months=24),
        Maturity(name="three_year", months=36),
        Maturity(name="five_year", months=60),
        Maturity(name="seven_year", months=84),
        Maturity(name="ten_year", months=120),
        Maturity(name="twenty_year", months=240),
        Maturity(name="thirty_year", months=360),
    )

    date = models.DateField(null=False, blank=False)

    one_month = models.FloatField(null=True, blank=False, editable=False)
    two_month = models.FloatField(null=True, blank=False, editable=False)
    three_month = models.FloatField(null=True, blank=False, editable=False)
    four_month = models.FloatField(null=True, blank=False, editable=False)
    six_month = models.FloatField(null=True, blank=False, editable=False)

    one_year = models.FloatField(null=False, blank=False)
    two_year = models.FloatField(null=False, blank=False)
    three_year = models.FloatField(null=False, blank=False)
    five_year = models.FloatField(null=False, blank=False)
    seven_year = models.FloatField(null=False, blank=False)
    ten_year = models.FloatField(null=False, blank=False)
    twenty_year = models.FloatField(null=False, blank=False)
    thirty_year = models.FloatField(null=False, blank=False)

    chart = models.ImageField(
        null=True,
        blank=True,
        editable=False,
        upload_to="uploads/%Y/%m/%d/",
    )

    class Meta:
        constraints = [
            models.UniqueConstraint("date", name="unique_date"),
        ]
        verbose_name_plural = "Treasury Data"

    def __str__(self):
        return f"<{self.__class__.__name__} {self.date}>"

    def __repr__(self):
        return str(self)

    def image_tag(self):
        return mark_safe(f'<img src="{self.chart.url}" width="600" height="400" />')

    image_tag.short_description = "Chart"

    @classmethod
    def _effective_maturities(cls, maturity):
        for m in cls._maturity_order:
            if m.months < maturity.months:
                yield m

    def _retrieve_treasury_data(self):
        current_date = timezone.now().date()

        if (
            current_date.month == self.date.month
            and current_date.year == self.date.year
        ):
            URL = CURRENT_MONTH_TREASURY_URL_TEMPLATE.format(
                year=self.date.year, month=f"{self.date.month:02}"
            )
        else:
            URL = ANNUAL_TREASURY_URL_TEMPLATE.format(year=self.date.year)

        resp = requests.get(URL, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()

        data = StringIO(resp.content.decode("utf-8"))
        dict_reader = csv.DictReader(data)

        for row in dict_reader:
            data_date = parse(row.pop("Date"))

            if data_date.date() == self.date:
                for key, val in row.items():
                    try:
                        if val != "":
                            val = float(val)
                        else:
                            val = None
                        setattr(self, self._treasury_map[key], val)
                    except Exception as e:
                        raise Exception(
                            f"Got error getting {key} treasury data for {self.date}: {e}"
                        )
                break
        else:
            raise Exception(f"Treasury data for {self.date} was not found")

        self.generate_chart()

    def save(self, *args, **kwargs):
        self._retrieve_treasury_data()
        super().save(*args, **kwargs)

    def par_values(self):
        return [getattr(self, maturity.name) for maturity in self._maturity_order]

    def zero_rates(self):
        zrs = {}

        for idx, maturity in enumerate(self._maturity_order):
            par_rate = getattr(self, maturity.name) / 100
            df = 1 / ((1 + (par_rate / 2)) ** (maturity.months / 6))
            zero_rate = 1 / (df ** (1 / maturity.months)) - 1
            zrs[maturity] = zero_rate
        return zrs

    def get_row_data(self):
        key_data = {}
        zero_rates = self.zero_rates()

        for idx, maturity in enumerate(self._maturity_order):
            months = maturity.months
            par = getattr(self, maturity.name) / 100
            zero_rate = zero_rates[maturity]
            zero = (1 + zero_rate) ** 12 - 1
            df = 1 / (1 + zero_rate) ** months
            ln_df = math.log(df)

            key_data[maturity.months] = DataRow(
                months=months,
                par=par,
                zero=zero,
                zero_rate=zero_rate,
                df=df,
                ln_df=ln_df,
            )

        key_slopes = []
        key_data_list = list(key_data.values())
        for idx in range(1, len(key_data)):
            curr = key_data_list[idx]
            prev = key_data_list[idx - 1]

            slope = (curr.ln_df - prev.ln_df) / (curr.months - prev.months)
            key_slopes.append(slope)

        data = []
        key_row = None
        key_slope = None
        for months in range(12, 361):
            if months in key_data:
                key_row = key_data[months]
                if key_slopes:
                    key_slope = key_slopes.pop(0)
                data.append(key_row)
                continue

            ln_df = key_row.ln_df + key_slope * (months - key_row.months)
            df = math.exp(ln_df)
            par = ""
            zero_rate = (1 / df) ** (1 / months) - 1
            zero = (1 + zero_rate) ** 12 - 1

            data.append(
                DataRow(
                    months=months,
                    par=par,
                    zero=zero,
                    zero_rate=zero_rate,
                    df=df,
                    ln_df=ln_df,
                )
            )
        return data

    def to_csv(self):
        output = StringIO()
        csv_writer = csv.writer(output)

        csv_writer.writerow(
            [
                "Months",
                "Par",
                "Zero",
                "",
                "DF",
                "ln(DF)",
            ]
        )

        for data in self.get_row_data():
            csv_writer.writerow(
                [
                    data.months,
                    data.par,
                    data.zero,
                    data.zero_rate,
                    data.df,
                    data.ln_df,
                ]
            )
        output.seek(0)
        return output

    def generate_chart(self):
        io = BytesIO()

        data = self.get_row_data()
        x_values = [x.months for x in data]
        y_values = [y.zero * 100 for y in data]
        plt.scatter(x_values, y_values)
        plt.title("Continuous Monthly Zero Rates")
        plt.xlabel("Months")
        plt.ylabel("Interest Rates (%)")
        plt.grid(True)
        plt.savefig(io)
        self.chart.save(
            f"{self.date.year}-{self.date.month}-{self.date.day}.png",
            ContentFile(io.getvalue()),
            save=False,
        )
        plt.close()


class Excel:
    def __init__(self):
        self.wb = load_workbook(filename=EXCEL_TEMPLATE_FILE)
        self.ws = None

    def populate_data(self, par_values, zero_rates):
        if self.ws is None:
            raise Exception("Sheet must be added before populating data")

        for idx, par_val in enumerate(par_values, start=PAR_VALUES_START_ROW):
            self.ws[f"{PAR_VALUES_COL}{idx}"] = par_val / 100

        for idx, zero_rate in enumerate(
            list(zero_rates.values()), start=ZERO_RATES_START_ROW
        ):
            self.ws[f"{ZERO_RATES_COL}{idx}"] = zero_rate

        chart = ScatterChart()
        chart.x_axis.title = "Future Monthly Periods"
        chart.y_axis.title = "Interest Rate"

        par_x_values = Reference(
            self.ws,
            min_col=MINI_CHART_MONTH_COL_INDEX,
            min_row=MINI_CHART_MIN_ROW,
            max_row=MINI_CHART_MAX_ROW,
        )

        par_values = Reference(
            self.ws,
            min_col=CHART_PAR_COL,
            min_row=MINI_CHART_MIN_ROW,
            max_row=MINI_CHART_MAX_ROW,
        )
        par_series = Series(par_values, par_x_values, title="Par")

        zero_x_values = Reference(
            self.ws,
            min_col=CHART_MONTH_COL_INDEX,
            min_row=CHART_MIN_ROW,
            max_row=CHART_MAX_ROW,
        )
        zero_values = Reference(
            self.ws,
            min_col=CHART_ZERO_COL,
            min_row=CHART_MIN_ROW,
            max_row=CHART_MAX_ROW,
        )
        zero_series = Series(zero_values, zero_x_values, title="Zero")

        chart.series.append(par_series)
        chart.series.append(zero_series)

        self.ws.add_chart(chart, "A2")

    def add_sheet(self, treasury_data):
        if self.ws is None:
            self.ws = self.wb.active
        else:
            source = self.ws
            self.ws = self.wb.copy_worksheet(source)

        self.ws.sheet_view.showGridLines = False
        self.ws.title = f"{treasury_data.date}"
        self.populate_data(treasury_data.par_values(), treasury_data.zero_rates())

    def stream(self):
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output
